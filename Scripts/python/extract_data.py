from openpyxl import load_workbook
import os
from shutil import copyfile
from datetime import datetime
from utils import get_first_row_free

# Lista dos arquivos txt
lista_arquivos = os.listdir("..\\..\\Data\\Extratos\\Caixa\\DGI\\")
template = "..\\..\\Data\\Template\\Template.xlsx"

for arquivo in lista_arquivos:
    print(arquivo)
    # Inicializa variaveis
    lista_empregados = []
    nome_empresa = ""
    cod_empresa = ""
    data_extrato = ""

    # Inicia leitura do arquivo.txt
    with open("..\\..\\Data\\Extratos\\Caixa\\DGI\\" + arquivo, "r+") as f:
        lines = f.readlines()
        lines = lines[2:-2]

        for idx, line in enumerate(lines):
            lista_pivo = []

            # Captura: RECLAMADA
            if line.upper().__contains__("NOME DA EMPRESA:"):
                if nome_empresa == "":
                    nome_empresa = line.split("NOME DA EMPRESA:")[1].strip()
                    nome_empresa = nome_empresa.replace("/", "")

            # Captura: CODIGO EMPRESA
            if line.upper().__contains__("CODIGO DA EMPRESA"):
                if cod_empresa == "":
                    cod_empresa = lines[idx + 1].split()[0].strip()
            
            # Captura: DATA DO EXTRATO
            if line.upper().__contains__("DATA CREDITO DE JAM"):
                if data_extrato == "":
                    data_extrato = lines[idx + 1].split()[4].strip()

            # Captura lista de Reclamantes
            if line.__contains__("VARA TRABALHISTA"):
                lista = lines[idx - 1].split("  ")
                for i in lista:
                    i.strip()
                    if len(i) > 1:
                        lista_pivo.append(i.strip())
                for i2 in line.split():
                    if i2.isdigit():
                        lista_pivo.append(i2.strip())
                lista_empregados.append(lista_pivo)
    
    # Verifica se o Relatório da empresa já criado
    if nome_empresa:
        relatorio = f"..\\..\\Data\\reports\\{nome_empresa}.xlsx"
        if os.path.exists(relatorio):
            index_excel = get_first_row_free(relatorio) - 1
            wb = load_workbook(relatorio, False)
            ws = wb.active
        else:
            index_excel = 1
            wb = load_workbook(template, False)
            ws = wb.active

        # Preeche Relatório
        if len(lista_empregados) > 0:
            for n in range(len(lista_empregados)):
                index_excel += 1

                # VARA
                ws["AS{}".format(index_excel)] = lista_empregados[n][8]

                # PROCESSO LOCALIZADO
                ws["AT{}".format(index_excel)] = lista_empregados[n][9]

                # RECLAMANTE
                ws["F{}".format(index_excel)] = lista_empregados[n][0]

                # RECLAMADA
                ws["E{}".format(index_excel)] = nome_empresa

                # DATA DE EXTRATO
                ws["H{}".format(index_excel)] = data_extrato

                # TIPO DE DEPOSITO
                ws["K{}".format(index_excel)] = "DGI"

                # CODIGO EMPRESA
                ws["AF{}".format(index_excel)] = cod_empresa

                # CODIGO EMPREGADO
                ws["AG{}".format(index_excel)] = lista_empregados[n][1]

                # DATA ADMISSAO
                ws["AH{}".format(index_excel)] = lista_empregados[n][2]

                # PIS
                ws["AJ{}".format(index_excel)] = lista_empregados[n][3]

                # VALOR
                ws["G{}".format(index_excel)] = lista_empregados[n][6]

            wb.save(relatorio)
        wb.close()
