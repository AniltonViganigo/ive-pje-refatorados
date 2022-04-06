from openpyxl import load_workbook

def get_first_row_free(file_path):
    wb = load_workbook(file_path, True)
    ws = wb.active

    cell_free = False
    
    counter = 0
    while cell_free == False:
        counter += 1
        current_cell = ws[f"D{counter}"].value
        if current_cell is None: 
            cell_free = True
            wb.close()
            return counter